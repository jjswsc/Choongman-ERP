# -*- coding: utf-8 -*-
"""Read cost Excel file structure"""
import os
import sys

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

base = os.path.dirname(os.path.abspath(__file__))
path = os.path.join(base, "..", "vercel-app", "원가분석.xlsx")
path = os.path.normpath(path)

if not os.path.exists(path):
    print("File not found:", path)
    sys.exit(1)

out_path = os.path.normpath(os.path.join(base, "..", "vercel-app", "cost_excel_structure.txt"))
try:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
except Exception as ex:
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("ERROR: " + str(ex))
    sys.exit(1)
with open(out_path, "w", encoding="utf-8") as f:
    f.write("Reading: " + path + "\n\n")
    for sheet_name in wb.sheetnames:
        f.write(f"\n=== Sheet: {sheet_name} ===\n")
        ws = wb[sheet_name]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 50:
                f.write(str(row) + "\n")
            else:
                f.write("... (more rows)\n")
                break
wb.close()
