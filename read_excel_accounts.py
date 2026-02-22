# -*- coding: utf-8 -*-
import sys
path = r"c:\Users\S&J\OneDrive\Desktop\회계\Ekkamai\3. Income Statement - EKKAMAI - Sep. 2025.xlsx"
try:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"=== Sheet: {sheet} ===")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 60:
                print(row)
        print()
    wb.close()
except ImportError:
    try:
        import xlrd
        wb = xlrd.open_workbook(path)
        for i, sheet in enumerate(wb.sheets()):
            print(f"=== Sheet: {sheet.name} ===")
            for r in range(min(60, sheet.nrows)):
                print(sheet.row_values(r))
            print()
    except Exception as e:
        print("xlrd Error:", e)
        sys.exit(1)
except Exception as e:
    print("Error:", e)
    sys.exit(1)
